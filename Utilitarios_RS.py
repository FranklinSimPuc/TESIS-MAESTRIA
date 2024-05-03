import random as rand
from math import e, trunc
import Clase_OpenDSS as OPEND
from matplotlib import pyplot as plt
import openpyxl
from statistics import mean,pstdev

rand.seed()

def SolAle(Num,Mallas):
    Sol=[0]*Num
    tamanho=len(Mallas)
    for i in range(Num):
        Vect_elemento=[0]*tamanho
        for j in range(tamanho):
            elem=rand.choice(Mallas[j])
            Vect_elemento[j]=elem
        Sol[i]=Vect_elemento
    return Sol

def Vecina1(X,Mallas):
    tam=len(X)
    Valor_X=X.copy()
    Vecina=X.copy()
    #Cantidad=abs(rand.uniform(-(tam-3),(tam-3))) #Cantidad de posiciones que cambiaran de valor
    #print(Cantidad)
    #Cantidad=int(Cantidad+1)
    Cantidad=3
    Valores=rand.sample(Valor_X,Cantidad).copy() #"sample"-Devuelve "n" elementos aleatorios de la baraja
    #print('elemento',Valores,Valor_X)
    for i in range(Cantidad):
        #print(i)
        elemento=Valores[i]
        #print(elemento)
        indice=Valor_X.index(elemento)
        #print('indice=',indice)
        #print(Mallas[indice])
        valor=Valor_X[indice]
        #print(valor)
        while Vecina[indice]==valor:
            Vecina[indice]=rand.choice(Mallas[indice])
    #print(Vecina)
    return Vecina

def Vecina2(X,Mallas):
    tam=len(X)
    Vecina=[0]*tam
    elemento1=rand.choice(X) 
    elemento2=rand.choice(X)
    indice1=X.index(elemento1)
    indice2=X.index(elemento2)
    Vecina[0]=rand.choice(Mallas[0])
    Vecina[1]=rand.choice(Mallas[1])
    Vecina[2]=rand.choice(Mallas[2])
    Vecina[3]=rand.choice(Mallas[3])
    Vecina[4]=rand.choice(Mallas[4])
    Vecina[indice1]=elemento1
    Vecina[indice2]=elemento2
    return Vecina   

def Vecina(X,Mallas):
    tam=len(X)
    Vecina_Sol=X.copy()
    #N_aleatorio=rand.uniform(0,tam)
    N_aleatorio=int(round(abs(rand.uniform(-tam,tam)),0))
    #solución actual
    lim_inf=trunc(N_aleatorio)
    lim_sup=lim_inf+1
    
    if lim_inf<N_aleatorio<=lim_sup:
        Valor_XX=abs(rand.randint(-len(Mallas[lim_sup-1])+1,len(Mallas[lim_sup-1])-1))
        Vecina_Sol[lim_sup-1]=Mallas[lim_sup-1][Valor_XX]  
    elif N_aleatorio<tam:
        Valor_XX=abs(rand.randint(-len(Mallas[N_aleatorio])+1,len(Mallas[N_aleatorio])-1))
        Vecina_Sol[N_aleatorio]=Mallas[N_aleatorio][Valor_XX]
    else:
        N_aleatorio=rand.randint(0,tam-1)
        Valor_XX=abs(rand.randint(-len(Mallas[N_aleatorio])+1,len(Mallas[N_aleatorio])-1))
        Vecina_Sol[N_aleatorio]=Mallas[N_aleatorio][Valor_XX]         
    return Vecina_Sol

def Vecina_(X,Mallas):
    tam=len(X)
    Vecina_Sol=X.copy()
    N_aleatorio=rand.uniform(0,tam)
    #N_aleatorio=int(round(abs(rand.uniform(-tam,tam)),0))
    #solución actual
    lim_inf=trunc(N_aleatorio)
    lim_sup=lim_inf+1
 
    if lim_inf<N_aleatorio<=lim_sup:
        Valor_XX=abs(rand.randint(-len(Mallas[lim_sup-1])+1,len(Mallas[lim_sup-1])-1))
        Vecina_Sol[lim_sup-1]=Mallas[lim_sup-1][Valor_XX]  
    elif N_aleatorio<tam:
        Valor_XX=abs(rand.randint(-len(Mallas[N_aleatorio])+1,len(Mallas[N_aleatorio])-1))
        Vecina_Sol[N_aleatorio]=Mallas[N_aleatorio][Valor_XX]
    else:
        N_aleatorio=rand.randint(0,tam-1)
        Valor_XX=abs(rand.randint(-len(Mallas[N_aleatorio])+1,len(Mallas[N_aleatorio])-1))
        Vecina_Sol[N_aleatorio]=Mallas[N_aleatorio][Valor_XX]         
    return Vecina_Sol

def Vecina_ISA(X,Mallas):
    tam=len(Mallas)-1
    Vecina_Sol=X.copy()
    m=rand.randint(1,3)
    ran=trunc(tam/2)

    if m==1:
        N_aleatorio=int(round(abs(rand.uniform(-tam,tam)),0))
    elif m==2:
        N_aleatorio=int(round(abs(rand.uniform(0,ran)),0))
    else:
        N_aleatorio=int(round(abs(rand.uniform(ran,tam)),0))        
    
    malla=Mallas[N_aleatorio]
    val=rand.choice(malla)
    if val!=Vecina_Sol[N_aleatorio]:
        Vecina_Sol[N_aleatorio]=val
    else:
        while val==Vecina_Sol[N_aleatorio]:
            val=rand.choice(malla)
        Vecina_Sol[N_aleatorio]=val
    return Vecina_Sol


def SA_Mejorado(To,Tf,max_vecinos,max_RepFO,alpha1,alpha2,Klm,Tlm,Solucion_Ini,Perdidas_Ini,Mallas,Sistema):

    if (Sistema==5 or Sistema=='5'):
        Objeto=OPEND.DSS(r"[D:\codigos\5Barras.dss]")
    elif (Sistema==16 or Sistema=='16'):
        Objeto=OPEND.DSS(r"[D:\codigos\16_barras.dss]")
    elif (Sistema==33 or Sistema=='33'):
        Objeto=OPEND.DSS(r"[D:\codigos\33Barras.dss]")
    elif (Sistema==69 or Sistema=='69'):
        Objeto=OPEND.DSS(r"[D:\codigos\69_barras.dss]") 
    elif (Sistema=='70a'):
        Objeto=OPEND.DSS(r"[D:\codigos\70a_barras.dss]") 
    elif (Sistema==70 or Sistema=='70'):
        Objeto=OPEND.DSS(r"[D:\codigos\70_barras.dss]") 
    elif (Sistema==94 or Sistema=='94'):
        Objeto=OPEND.DSS(r"[D:\codigos\94Barras.dss]")
    elif (Sistema==135 or Sistema=='135'):
        Objeto=OPEND.DSS(r"[D:\codigos\135Barras.dss]")
    Objeto.compile_DSS()
    
    T=To
    dT=To-Tf

    iter=[]
    Vector_T=[]
    Vector_FO=[]
    Vector_FOact=[]

    #Solución actual
    Valor_X=Solucion_Ini.copy()
    Perdida_act=Perdidas_Ini

    #Mejor solucion
    Valor_Xm=Valor_X.copy()
    Perdida_mej=Perdida_act

    k=1
    exitos=10000
    repFOtem=0
    MLD=0 #iteraciones para Lundy & Mees

    Vector_T.append(T)
    Vector_FO.append(Perdida_mej)
    Vector_FOact.append(Perdida_act)
    iter.append(0)  

    #while (T>=Tf) and (exitos!=0) and (repFOtem<max_RepFO):
    while (T>=Tf) and (repFOtem<max_RepFO):
        exitos=0
        vecinos=0
        #print(T)
        while exitos<=round(0.1*max_vecinos) and vecinos<=max_vecinos:
            #Creacion de vecino         
            Valor_XX=Vecina_ISA(Valor_X,Mallas).copy()
            TenMin,Perdida_cand=Objeto.FitnessVecino(Valor_XX,'0')
            #print(Valor_XX,Perdida_cand)
            delta= Perdida_cand-Perdida_act
            if (delta<0 and TenMin>0.0001):
                Valor_X=Valor_XX.copy()
                Perdida_act=Perdida_cand
                exitos=exitos+1

                if (Perdida_cand<Perdida_mej):
                    Valor_Xm=Valor_XX.copy()
                    Perdida_mej=Perdida_cand
                    repFOtem=0

            else:
                Boolsman=round(e**(-delta/T),10)

                if (rand.random()<Boolsman):
                    Valor_X=Valor_XX.copy()
                    Perdida_act=Perdida_cand
                    #exitos=exitos+1

            vecinos=vecinos+1
        
        if (Perdida_act==Perdida_mej):
            repFOtem=repFOtem+1

        if (T>To-0.3*dT) and (MLD==0):
            T=alpha1*T
        elif (To-0.8*dT<T<=To-0.3*dT) and (MLD==0):
            T=alpha2*T
        else:
            if(MLD==0):
                MLD=Klm*k # en defecto es 2 multiplica para que exista el doble de iteraciones que se se tiene hasta ahora
                beta=(dT/(MLD*Tf*To))
                #T=To/(1+beta*To)
                T=Tlm*To
            else:
                T=T/(1+beta*T)
        
        Vector_T.append(T)
        Vector_FO.append(Perdida_mej)
        Vector_FOact.append(Perdida_act)
        iter.append(k)
        k=k+1
    
    print("Repeticiones --> ",repFOtem,"---","Exitos--> ",exitos,"---","Temperatura final --> ",T)
    return Valor_Xm,Perdida_mej,iter,Vector_T,Vector_FO,Vector_FOact


def SA_Clasico(To,Tf,max_vecinos,alpha1,Solucion_Ini,Perdidas_Ini,Mallas,Sistema):

    if (Sistema==5 or Sistema=='5'):
        Objeto=OPEND.DSS(r"[D:\codigos\5Barras.dss]")
    elif (Sistema==16 or Sistema=='16'):
        Objeto=OPEND.DSS(r"[D:\codigos\16_barras.dss]")
    elif (Sistema==33 or Sistema=='33'):
        Objeto=OPEND.DSS(r"[D:\codigos\33Barras.dss]")
    elif (Sistema==69 or Sistema=='69'):
        Objeto=OPEND.DSS(r"[D:\codigos\69_barras.dss]") 
    elif (Sistema=='70a'):
        Objeto=OPEND.DSS(r"[D:\codigos\70a_barras.dss]") 
    elif (Sistema==70 or Sistema=='70'):
        Objeto=OPEND.DSS(r"[D:\codigos\70_barras.dss]") 
    elif (Sistema==94 or Sistema=='94'):
        Objeto=OPEND.DSS(r"[D:\codigos\94Barras.dss]")
    elif (Sistema==135 or Sistema=='135'):
        Objeto=OPEND.DSS(r"[D:\codigos\135Barras.dss]")
    Objeto.compile_DSS()
    
    T=To

    iter=[]
    Vector_T=[]
    Vector_FO=[]
    Vector_FOact=[]

    #Solución actual
    Valor_X=Solucion_Ini.copy()
    Perdida_act=Perdidas_Ini

    #Mejor solucion
    Valor_Xm=Valor_X.copy()
    Perdida_mej=Perdida_act

    k=1

    Vector_T.append(T)
    Vector_FO.append(Perdida_mej)
    Vector_FOact.append(Perdida_act)
    iter.append(0)  

    while (T>=Tf):
        vecinos=0
        #print(T)
        while vecinos<=max_vecinos:
            #Creacion de vecino         
            Valor_XX=Vecina_(Valor_X,Mallas).copy()
            TenMin,Perdida_cand=Objeto.FitnessVecino(Valor_XX,'0')
            #print(Valor_XX,Perdida_cand)
            delta= Perdida_cand-Perdida_act
            if (delta<0 and TenMin>0.0001):
                Valor_X=Valor_XX.copy()
                Perdida_act=Perdida_cand

                if (Perdida_cand<Perdida_mej):
                    Valor_Xm=Valor_XX.copy()
                    Perdida_mej=Perdida_cand

            else:
                Boolsman=round(e**(-delta/T),10)

                if (rand.random()<Boolsman):
                    Valor_X=Valor_XX.copy()
                    Perdida_act=Perdida_cand

            vecinos=vecinos+1
        

        T=alpha1*T
        
        Vector_T.append(T)
        Vector_FO.append(Perdida_mej)
        Vector_FOact.append(Perdida_act)
        iter.append(k)
        k=k+1
    
    print("Temperatura final --> ",T)
    return Valor_Xm,Perdida_mej,iter,Vector_T,Vector_FO,Vector_FOact

def SA_joao(To,Tf,max_vecinos,alpha1,Solucion_Ini,Perdidas_Ini,Mallas,Sistema):

    if (Sistema==5 or Sistema=='5'):
        Objeto=OPEND.DSS(r"[D:\codigos\5Barras.dss]")
    elif (Sistema==16 or Sistema=='16'):
        Objeto=OPEND.DSS(r"[D:\codigos\16_barras.dss]")
    elif (Sistema==33 or Sistema=='33'):
        Objeto=OPEND.DSS(r"[D:\codigos\33Barras.dss]")
    elif (Sistema==69 or Sistema=='69'):
        Objeto=OPEND.DSS(r"[D:\codigos\69_barras.dss]") 
    elif (Sistema=='70a'):
        Objeto=OPEND.DSS(r"[D:\codigos\70a_barras.dss]") 
    elif (Sistema==70 or Sistema=='70'):
        Objeto=OPEND.DSS(r"[D:\codigos\70_barras.dss]") 
    elif (Sistema==94 or Sistema=='94'):
        Objeto=OPEND.DSS(r"[D:\codigos\94Barras.dss]")
    elif (Sistema==135 or Sistema=='135'):
        Objeto=OPEND.DSS(r"[D:\codigos\135Barras.dss]")
    Objeto.compile_DSS()
    
    T=To

    iter=[]
    Vector_T=[]
    Vector_FO=[]
    Vector_FOact=[]

    #Solución actual
    Valor_X=Solucion_Ini.copy()
    Perdida_act=Perdidas_Ini

    #Mejor solucion
    Valor_Xm=Valor_X.copy()
    Perdida_mej=Perdida_act

    k=1

    Vector_T.append(T)
    Vector_FO.append(Perdida_mej)
    Vector_FOact.append(Perdida_act)
    iter.append(0)  

    while (T>=Tf):

        vecinos=0
        #print(T)
        while vecinos<=max_vecinos:
            #Creacion de vecino         
            Valor_XX=Vecina_(Valor_X,Mallas).copy()
            TenMin,Perdida_cand=Objeto.FitnessVecino(Valor_XX,'0')
            #print(Valor_XX,Perdida_cand)
            delta= Perdida_cand-Perdida_act
            if (delta<0 and TenMin>0.0001):
                Valor_X=Valor_XX.copy()
                Perdida_act=Perdida_cand

                if (Perdida_cand<Perdida_mej):
                    Valor_Xm=Valor_XX.copy()
                    Perdida_mej=Perdida_cand

            else:
                Boolsman=round(e**(-delta/T),10)

                if (rand.random()<Boolsman):
                    Valor_X=Valor_XX.copy()
                    Perdida_act=Perdida_cand

            vecinos=vecinos+1
        

        T=alpha1*T
        
        Vector_T.append(T)
        Vector_FO.append(Perdida_mej)
        Vector_FOact.append(Perdida_act)
        iter.append(k)
        k=k+1
    
    print("Temperatura final --> ",T)
    return Valor_Xm,Perdida_mej,iter,Vector_T,Vector_FO,Vector_FOact

def plot(X,Y):
    ax=plt.gca()
    ax.plot(X,Y,'r')
    return (ax)

def plotcomp(X1,Y1,X2,Y2,X3,Y3):
    ax=plt.gca()
    ax.plot(X1,Y1,'r',label='ISA-HC')
    ax.plot(X2,Y2,'--b',label='Joao Amorim')
    ax.plot(X3,Y3,'--g',label='Andrade')
    return (ax)

def plotcomp(X1,Y1,X2,Y2,X3,Y3,X4,Y4):
    ax=plt.gca()
    ax.plot(X1,Y1,'--r',label='ISA-HC C=0.1')
    ax.plot(X2,Y2,'--b',label='Joao Amorim C=0.8')
    ax.plot(X3,Y3,'--g',label='Andrade C=0.1')
    ax.plot(X4,Y4,'--c',label='Andrade C=0.8')
    return (ax)

def Escribir_xls_1corrida(archivo,iter,Vector_T,Vector_FO,Vector_FOact):
    wb=openpyxl.Workbook()
    tam=len(iter)
    sheet=wb.create_sheet(index=1,title='Solucion')
    c=sheet.cell(row=1,column=1)
    c.value='iter'
    c=sheet.cell(row=1,column=2)
    c.value='T'
    cc=sheet.cell(row=1,column=3)
    cc.value='FOm' 
    cc=sheet.cell(row=1,column=4)
    cc.value='FOact'

    for i in range(tam):
        c0=sheet.cell(row=i+2,column=1)
        c0.value=iter[i]      
        c0=sheet.cell(row=i+2,column=2)
        c0.value=Vector_T[i]  
        c1=sheet.cell(row=i+2,column=3)
        c1.value=Vector_FO[i]
        c2=sheet.cell(row=i+2,column=4)
        c2.value=Vector_FOact[i]

    wb.save(archivo)

def Escribir_xls(archivo,X,To,X1,Y1,Tf,XX1,YY1,tiempo,Vglobal_iter,Vglobal_T,Vglobal_FOm,Vglobal_FOact):
    tamano=len(X)
    wb=openpyxl.Workbook()
    sheet=wb.active
    sheet=wb.create_sheet(index=1,title='Resultados')
    c=sheet.cell(row=1,column=1)
    c.value='N°_Sim'
    c=sheet.cell(row=1,column=2)
    c.value='To'
    cc=sheet.cell(row=1,column=3)
    cc.value='Solución Inicial' 
    cc=sheet.cell(row=1,column=4)
    cc.value='Pérdida Inicial kW'
    c=sheet.cell(row=1,column=5)
    c.value='Tf'
    cc=sheet.cell(row=1,column=6)
    cc.value='Solución Final'
    cc=sheet.cell(row=1,column=7)
    cc.value='Pérdida Final kW'

    for i in range(tamano):
        c0=sheet.cell(row=i+2,column=1)
        c0.value=X[i]      
        c0=sheet.cell(row=i+2,column=2)
        c0.value=To[i]  
        c1=sheet.cell(row=i+2,column=3)
        c1.value=str(X1[i])
        c2=sheet.cell(row=i+2,column=4)
        c2.value=Y1[i]
        c0=sheet.cell(row=i+2,column=5)
        c0.value=Tf[i] 
        c3=sheet.cell(row=i+2,column=6)
        c3.value=str(XX1[i])
        c4=sheet.cell(row=i+2,column=7)
        c4.value=YY1[i]

    sheet=wb.create_sheet(index=1,title='Resumen')
    cc=sheet.cell(row=1,column=1)
    cc.value='Tiempo total seg'
    dd=sheet.cell(row=2,column=1)
    dd.value=tiempo
    cc=sheet.cell(row=1,column=2)
    cc.value='Perdida Promedio'
    dd=sheet.cell(row=2,column=2)
    dd.value=mean(YY1)
    cc=sheet.cell(row=1,column=3)
    cc.value='Desv. Estandar'
    dd=sheet.cell(row=2,column=3)
    dd.value=pstdev(YY1)
    cc=sheet.cell(row=1,column=4)
    cc.value='Peor Perdida'
    dd=sheet.cell(row=2,column=4)
    val=max(YY1)
    dd.value=val
    indice=YY1.index(val)
    cc=sheet.cell(row=1,column=5)
    cc.value='Peor Solucion'
    dd=sheet.cell(row=2,column=5)
    dd.value=str(XX1[indice])
    cc=sheet.cell(row=1,column=6)
    cc.value='Mejor Perdida'
    dd=sheet.cell(row=2,column=6)
    val=min(YY1)
    dd.value=val
    indice=YY1.index(val)
    cc=sheet.cell(row=1,column=7)
    cc.value='Mejor Solucion'
    dd=sheet.cell(row=2,column=7)
    solucion=str(XX1[indice])
    dd.value=solucion
    cc=sheet.cell(row=1,column=8)
    cc.value='Nro Recurrencias'
    dd=sheet.cell(row=2,column=8)
    dd.value=XX1.count(solucion)

    tam=len(Vglobal_iter)
    sheet=wb.create_sheet(index=1,title='MejorSolucion')
    c=sheet.cell(row=1,column=1)
    c.value='iter'
    c=sheet.cell(row=1,column=2)
    c.value='T'
    cc=sheet.cell(row=1,column=3)
    cc.value='FOm' 
    cc=sheet.cell(row=1,column=4)
    cc.value='FOact'

    for i in range(tam):
        c0=sheet.cell(row=i+2,column=1)
        c0.value=Vglobal_iter[i]      
        c0=sheet.cell(row=i+2,column=2)
        c0.value=Vglobal_T[i]  
        c1=sheet.cell(row=i+2,column=3)
        c1.value=Vglobal_FOm[i]
        c2=sheet.cell(row=i+2,column=4)
        c2.value=Vglobal_FOact[i]

    wb.save(archivo)